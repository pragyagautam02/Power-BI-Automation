## imported necessary modules
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.alignment import Alignment
import codecs
import json
import zipfile
import pathlib
from pathlib import Path
from dax_extract import read_data_model_schema
import requests
import openai
import os
import tkinter
from tkinter import filedialog
from tkinter.filedialog import askdirectory
import xlwt
from xlwt import Workbook
import pandas as pd
from xlwt import Pattern
import shutil

## created open AI key
openai.api_key = "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"


## function defined to extract the data in xls format ----------------------------------------------------------------------------------------------------------------
def xls_extract(data, file, base_file_name, json_path):
    ## MEASURES ----------------------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title="Measures")
    # ws.append(['Measure Name','Measure Expression','Measure Data Type','Measure Description'])
    cnt = 1
    for t in data['model']['tables']:
        if 'measures' in t:
            list_measures = t['measures']
            for i in list_measures:
                prompt = "Explain the following calculation in a few sentences in simple business terms without using DAX function names: " + \
                         i['expression']
                completion = openai.Completion.create(engine="text-davinci-003", prompt=prompt, max_tokens=64)
                #
                # print(completion.choices[0]['text'].strip())
                i['description'] = completion.choices[0]['text'].strip()
                cnt += 1
                if (cnt == 2): ws.append(
                    ['Measure Name', 'Measure Expression', 'Measure Data Type', 'Measure Description'])
                ws.append([i['name'], i['expression'], i['dataType'], i['description']])

    with codecs.open(json_path, 'w', 'utf-16-le') as f:
        json.dump(data, f, indent=4)
    if (cnt >= 2):
        table = Table(displayName="Table1", ref="A1:D" + str(cnt))
        ws.add_table(table)
        # Apply some style to the table
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        for col in ws.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 50
    else:
        ws.append(['No measures present in this file'])
        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    # Measure the length of the cell value
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                adjusted_width = (max_length + 2) * 1.2  # Add some extra padding and multiply by a factor to account for different fonts and styles
                ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        # Wrap text in the cells
        for cell in column_cells:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    ## SOURCE INFORMATION ----------------------------------------------------------------------
    source = wb.create_sheet('Source Information')
    i = 0
    if 'tables' in data['model']:
        source.append(['Table No', 'Table Name', 'Table Type', 'Table Source','Original Table Name', 'Table Query', 'Modification','Modification Description'])

        for t in data['model']['tables']:
            if 'partitions' in t:
                list_partitions = t['partitions']
                List_source = (list_partitions[0]['source']['expression'])
                if List_source[0] == 'let':
                    name = list_partitions[0]['name'].split('-')[0]
                    i += 1
                    p = List_source[1]
                    # print(p)
                    Ttype = p.split("(")[0].split('= ')[1]

                    if '"' in p:
                        st = 0
                        ed = len(p) - 1
                        while (p[st] != '"'):
                            st += 1
                        while (p[ed] != '"'):
                            ed -= 1

                        TSource = p[st: ed + 1]
                        if "Query=" in TSource:
                            TSource = TSource.split("[Query")[0]
                        if "Delimiter=" in TSource:
                            TSource=TSource.split("),[Delimiter")[0]
                    else:
                        TSource = List_source[2]

                    otname = ""
                    if "Query=" in p:
                        otname = p.split("FROM")[1].split("#")[0]
                        otname = otname.replace("(lf)", "")
                        # print(TName)
                    elif "Sql." in p:
                        otname = List_source[2].split("=")[0].strip()
                        # print(TName)
                    elif "Excel." in p:
                        otname = List_source[2].split("=")[0].strip()
                        otname = otname.replace("#", "")
                        # print(TName)
                    elif "Dataflows" in p:
                        otname = List_source[5].split("=")[0].split("\"")[1].strip()
                        # print(TName)
                    else:
                        otname = name

                    TQuery = ""
                    if "Query=" in p:
                        TQuery = p.split("Query=")[1]
                        TQuery = TQuery.replace("#(lf)", " ")
                    else:
                        TQuery = "No Query"

                    idx = -1
                    for i1 in range(2, len(List_source)):
                        if len(List_source[i1]) > 5 and List_source[i1][4] == '#':
                            idx = i1
                            break
                    Tmodification = ""
                    completion = "No Description"
                    if idx==-1:
                        Tmodification = "No Modification"
                    else:
                        # Tmodification = '\n\n'.join(List_source[idx:-2])
                        pr = 1
                        for id in range(idx, len(List_source)-2):
                            p1 = List_source[id].split("    ")[1]
                            Tmodification += str(pr) + ". " + p1 + '\n\n'
                            pr += 1
                        prompt = " ".join(List_source[2:])
                        prompt = "Explain this in normal terms: " + prompt
                        #print(prompt)
                        completion = openai.Completion.create(engine="text-davinci-003", prompt=prompt, max_tokens=2048)
                        #print(completion)

                    t = ""
                    if completion != 'No Description':
                        t = completion.choices[0]['text'].strip()
                    else:
                        t = completion
                    source.append([i, name, Ttype, TSource, otname, TQuery, Tmodification,t])

        i += 1
        table = Table(displayName="Source", ref="A1:H" + str(i))
        source.add_table(table)

        # Apply some style to the table
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        for col in source.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        source.column_dimensions["A"].width = 20
        source.column_dimensions["B"].width = 20
        source.column_dimensions["C"].width = 30
        source.column_dimensions["D"].width = 40
        source.column_dimensions["E"].width = 20
        source.column_dimensions["F"].width = 50
        source.column_dimensions["G"].width = 80
        source.column_dimensions["H"].width = 80
    else:
        source.append(['No Source present in this file'])
        for column_cells in source.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    # Measure the length of the cell value
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Add some extra padding and multiply by a factor to account for different fonts and styles
            source.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            # Wrap text in the cells
            for cell in column_cells:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    ## RELATIOPNSHIPS ----------------------------------------------------------------------
    relation = wb.create_sheet('Relationships')
    if 'relationships' in data['model']:
        # relation=wb.create_sheet('Relationships')
        cnt1 = 0
        # relation.append(['From Table','From Column','To Table','To Column','State','Direction','Cardinality'])
        for t in data['model']['relationships']:
            d = ""
            a = ""
            if "joinOnDateBehavior" not in t:
                cnt1 += 1
                if "crossFilteringBehavior" in t:
                    d = "Both Directional"
                else:
                    d = "Single Directional"

                if "toCardinality" in t:
                    if t['toCardinality'] == "one":
                        a = 'One to one (1:1)'
                    elif t['toCardinality'] == "many":
                        a = 'Many to many (*:*)'
                    else:
                        pass

                elif "fromCardinality" in t:
                    if t['fromCardinality'] == "one":
                        a = 'One to one (1:1)'
                    elif t['fromCardinality'] == "many":
                        a = 'Many to many (*:*)'
                    else:
                        pass
                else:
                    a = 'Many to one (*:1)'

                if (cnt1 == 1): relation.append(['From Table', 'From Column', 'To Table', 'To Column', 'State', 'Direction', 'Cardinality'])
                relation.append([t['fromTable'], t['fromColumn'], t['toTable'], t['toColumn'], t['state'], d, a])

        if (cnt1 >= 1):
            table = Table(displayName="Relationships", ref="A1:G" + str(cnt1 + 1))
            relation.add_table(table)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                                   showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style

            for col in relation.columns:
                for cell in col:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            relation.column_dimensions["A"].width = 40
            relation.column_dimensions["B"].width = 20
            relation.column_dimensions["C"].width = 40
            relation.column_dimensions["D"].width = 20
            relation.column_dimensions["E"].width = 20
            relation.column_dimensions["F"].width = 20
            relation.column_dimensions["G"].width = 20

        else:
            relation.append(['No relationships present in this file'])
            for column_cells in relation.columns:
                max_length = 0
                for cell in column_cells:
                    try:
                        # Measure the length of the cell value
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                    adjusted_width = (max_length + 2) * 1.2  # Add some extra padding and multiply by a factor to account for different fonts and styles
                    relation.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            # Wrap text in the cells
            for cell in column_cells:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    else:
        relation.append(['No relationships present in this file'])
        for column_cells in relation.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    # Measure the length of the cell value
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Add some extra padding and multiply by a factor to account for different fonts and styles
            relation.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        # Wrap text in the cells
        for cell in column_cells:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    dir_name = os.path.dirname(file)
    new_dir = pathlib.Path(dir_name, "EXCEL Output")
    new_dir.mkdir(parents=True, exist_ok=True)
    file_name = base_file_name + ".xlsx"
    save1 = str(new_dir) + "\\" + file_name
    wb.save(save1)
    workbook = openpyxl.load_workbook(save1)
    sheet_to_remove = workbook['Sheet']
    workbook.remove(sheet_to_remove)
    workbook.save(save1)
    print("Created :", file_name)


## function defined to extract the data in json format ---------------------------------------------------------------------------------------
def json_extract(file):
    temp_dir = 'temp'
    os.makedirs(temp_dir, exist_ok=True)
    dir_name = os.path.dirname(file)
    base_file_name = Path(file).stem

    x = list(file)
    x[-1] = 't'
    file = ''.join(x)

    pbit_path = Path(file)
    with zipfile.ZipFile(pbit_path, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)

    # Load the data model schema JSON file
    json_path = os.path.join(temp_dir, 'DataModelSchema')
    with codecs.open(json_path, 'r', 'utf-16-le') as f:
        contents = f.read()

    data = json.loads(contents)
    # data = read_data_model_schema(pbit_path)

    new_dir = pathlib.Path(dir_name, "JSON Output")
    new_dir.mkdir(parents=True, exist_ok=True)
    file_name = base_file_name + ".json"

    save1 = str(new_dir) + "\\" + file_name
    out_file = open(Path(save1), "w")

    json.dump(data, out_file, indent=6)
    out_file.close()

    xls_extract(data, file, base_file_name, json_path)
    with codecs.open(json_path, 'r', 'utf-16-le') as f:
        contents = f.read()

    data = json.loads(contents)
    # data = read_data_model_schema(pbit_path)

    # new_dir = pathlib.Path(dir_name, "JSON Output")
    # new_dir.mkdir(parents=True, exist_ok=True)
    file_name = base_file_name + ".json"

    save1 = str(new_dir) + "\\" + file_name
    out_file = open(Path(save1), "w")

    json.dump(data, out_file, indent=6)
    out_file.close()
    print("Created :", file_name)

    dir_name = os.path.dirname(file)
    new_dir = pathlib.Path(dir_name, "UPDATED pbit")
    new_dir.mkdir(parents=True, exist_ok=True)

    new_pbit_path = base_file_name + '_updated.pbit'
    save1 = str(new_dir) + "\\" + new_pbit_path
    print("Created :", new_pbit_path)
    print()
    with zipfile.ZipFile(save1, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                file_path = os.path.join(root, file)
                rel_path = os.path.relpath(file_path, temp_dir)
                zip_ref.write(file_path, rel_path)

    shutil.rmtree(temp_dir)


## Main ----------------------------------------------------------------------------------------------------------------------------------------------------
print("Do you want to select file or folder\n1. File\n2. Multiple Files\n3. Folder")
op = int(input("\nEnter Option : "))

if (op == 1):
    file = filedialog.askopenfilename(title="Select file")
    print("Currently Processing {" + str(Path(file).stem) + "}...")
    json_extract(file)
elif(op == 2):
    files = filedialog.askopenfilenames(title="Select files")
    for f in files:
        print("Currently Processing {" + str(Path(f).stem) + "}...")
        json_extract(f)
elif(op == 3):
    ## opening file dialog to select folder
    folder = askdirectory(title="Select folder")
    for f in os.listdir(folder):
        ## searching for file with specific file extension
        if f.endswith('.pbix'):
            print("Currently Processing {" + str(f) + "}...")
            # Create the filepath of particular file
            file_path = f"{folder}/{f}"
            json_extract(file_path)
else:
    print("Invalid Input")